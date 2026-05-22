"""
Scraper Template — Reference for the find-leads skill.

Each scraper script should follow this pattern:
1. Import dependencies
2. Define constants (headers, mappings)
3. Fetch data with rate limiting and error handling
4. Deduplicate results
5. Write to XLSX using the standard format

Run with:
    PYTHONUNBUFFERED=1 uv run --no-project --with httpx,beautifulsoup4,lxml,openpyxl python3 scrape_<source>.py
"""
import httpx
import time
import re
import sys
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from collections import Counter

# ── Always enable unbuffered output for progress tracking ──
sys.stdout.reconfigure(line_buffering=True)

# ── Standard headers to avoid blocks ──
HEADERS = {
    "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/json",
    "Accept-Language": "en-US,en;q=0.9",
}

# ── XML-illegal character sanitization ──
ILLEGAL_CHARS_RE = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')
def clean(s):
    if not s: return s
    return ILLEGAL_CHARS_RE.sub('', str(s))


def scrape_api(base_url: str, params: dict = None) -> list[dict]:
    """
    Template for API-based scraping.
    Adapt the response parsing to match the actual API structure.
    """
    all_items = []
    errors = 0

    with httpx.Client(headers=HEADERS, timeout=30, follow_redirects=True) as client:
        # Step 1: Test with one request to verify structure
        print(f"Testing API at {base_url}...")
        test_resp = client.get(base_url, params=params)
        print(f"Status: {test_resp.status_code}")
        if test_resp.status_code != 200:
            print(f"API returned {test_resp.status_code}, aborting")
            return []

        # Step 2: Parse response (ADAPT THIS to match actual API)
        data = test_resp.json()
        # Example: items = data.get("results", [])
        # Example: total = data.get("total", 0)
        # Example: next_page = data.get("next", None)

        # Step 3: Paginate if needed
        # while next_page:
        #     resp = client.get(next_page)
        #     ...
        #     time.sleep(0.3)

    return all_items


def scrape_html_directory(urls: list[tuple[str, str]], parse_page) -> list[dict]:
    """
    Template for HTML directory scraping.
    urls: list of (url, label) tuples
    parse_page: function(soup, label) -> list[dict]
    """
    all_items = []
    errors = 0

    with httpx.Client(headers=HEADERS, timeout=30, follow_redirects=True) as client:
        for i, (url, label) in enumerate(urls):
            try:
                resp = client.get(url)
                if resp.status_code == 200:
                    soup = BeautifulSoup(resp.text, "lxml")
                    items = parse_page(soup, label)
                    all_items.extend(items)
                else:
                    errors += 1
            except Exception as e:
                errors += 1

            if (i + 1) % 50 == 0:
                print(f"  [{i+1}/{len(urls)}] {label} — {len(all_items)} leads so far, {errors} errors")

            time.sleep(0.4)  # Rate limit

    print(f"\nTotal: {len(all_items)} leads, {errors} errors")
    return all_items


def scrape_wikipedia_list(list_url: str, country: str = "") -> list[dict]:
    """
    Template for Wikipedia list scraping.
    Handles both table-based and list-based formats.
    """
    items = []

    with httpx.Client(headers=HEADERS, timeout=30, follow_redirects=True) as client:
        resp = client.get(list_url)
        if resp.status_code != 200:
            print(f"Wikipedia returned {resp.status_code} for {list_url}")
            return []

        soup = BeautifulSoup(resp.text, "lxml")

        # Try tables first
        for table in soup.find_all("table", class_=re.compile(r"wikitable|sortable")):
            headers_row = table.find("tr")
            if not headers_row:
                continue
            col_headers = [th.get_text(strip=True).lower() for th in headers_row.find_all(["th", "td"])]

            for row in table.find_all("tr")[1:]:
                cells = row.find_all(["td", "th"])
                if len(cells) < 2:
                    continue
                row_data = {col_headers[i]: cells[i].get_text(strip=True)
                           for i in range(min(len(col_headers), len(cells)))}
                # Extract first link as name if available
                first_link = cells[0].find("a")
                name = first_link.get_text(strip=True) if first_link else cells[0].get_text(strip=True)

                if name and not name.startswith("List of"):
                    items.append({
                        "name": name,
                        "country": country,
                        **row_data,
                    })

        # Try lists if no tables found
        if not items:
            for ul in soup.find_all(["ul", "ol"]):
                for li in ul.find_all("li", recursive=False):
                    link = li.find("a")
                    if link:
                        name = link.get_text(strip=True)
                        if name and not name.startswith("List of") and len(name) > 2:
                            items.append({"name": name, "country": country})

    return items


def deduplicate(items: list[dict], keys: tuple = ("name", "country", "city")) -> list[dict]:
    """Deduplicate by normalized compound key."""
    seen = set()
    unique = []
    for item in items:
        key = tuple(str(item.get(k, "")).lower().strip() for k in keys)
        if key not in seen:
            seen.add(key)
            unique.append(item)
    print(f"Dedup: {len(items)} -> {len(unique)} ({len(items) - len(unique)} duplicates removed)")
    return unique


def write_xlsx(
    filename: str,
    sheet_name: str,
    columns: list[str],
    data: list[dict],
    header_color: str = "2E86AB",
    column_widths: dict[str, int] | None = None,
    country_key: str = "country",
):
    """Write leads to a formatted XLSX file with summary sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Headers
    for c, header in enumerate(columns, 1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # Data
    col_keys = [col.lower().replace(" ", "_").replace("/", "_") for col in columns]
    for r, row_data in enumerate(data, 2):
        for c, key in enumerate(col_keys, 1):
            val = row_data.get(key, "")
            cell = ws.cell(row=r, column=c, value=clean(val) if isinstance(val, str) else val)
            cell.border = thin_border

    # Formatting
    ws.freeze_panes = "A2"
    last_col = chr(64 + min(len(columns), 26))
    ws.auto_filter.ref = f"A1:{last_col}{len(data) + 1}"

    if column_widths:
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

    # Summary sheet
    ws2 = wb.create_sheet("Summary by Country")
    country_counts = Counter(clean(str(row.get(country_key, "Unknown"))) for row in data)
    ws2.cell(row=1, column=1, value="Country").font = header_font
    ws2.cell(row=1, column=1).fill = header_fill
    ws2.cell(row=1, column=2, value="Leads").font = header_font
    ws2.cell(row=1, column=2).fill = header_fill
    for r, (country, count) in enumerate(sorted(country_counts.items(), key=lambda x: -x[1]), 2):
        ws2.cell(row=r, column=1, value=country)
        ws2.cell(row=r, column=2, value=count)
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 12

    wb.save(filename)
    print(f"Saved {filename} — {len(data)} leads from {len(country_counts)} countries")


# ── Main execution template ──
if __name__ == "__main__":
    print("This is a template — adapt it for your specific source")
    print("See SKILL.md for the full scraping methodology")
