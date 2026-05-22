"""
XLSX Lead Export Template — Reference for the find-leads skill.

Usage:
    uv run --no-project --with openpyxl python3 xlsx_template.py

This is a reference template showing the standard XLSX output format.
Copy and adapt the relevant parts for each scraper script.
"""
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from collections import Counter

# ── String sanitization ──────────────────────────────────────────────
# openpyxl crashes on XML-illegal characters (common in scraped URLs)
ILLEGAL_CHARS_RE = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')

def clean(s):
    """Remove XML-illegal characters from a string for safe XLSX writing."""
    if not s:
        return s
    return ILLEGAL_CHARS_RE.sub('', str(s))


# ── Color palette for different sources ──────────────────────────────
COLORS = {
    "blue":    "2E86AB",
    "green":   "2D8B4E",
    "red":     "D4380D",
    "purple":  "7B2D8E",
    "orange":  "E8710A",
    "teal":    "0D8B8B",
    "navy":    "1B3A5C",
    "brown":   "8B4513",
}


def create_leads_xlsx(
    filename: str,
    sheet_name: str,
    columns: list[str],
    data: list[dict],
    color: str = "blue",
    country_field: str = "country",
    column_widths: dict[str, int] | None = None,
):
    """
    Create a styled XLSX file with leads data and a summary sheet.

    Args:
        filename: Output file path (e.g., "leads_wikipedia.xlsx")
        sheet_name: Name for the main data sheet
        columns: List of column headers (e.g., ["Name", "Country", "City"])
        data: List of dicts, each with keys matching lowercase column names
              or with keys provided in the same order as columns
        color: Header color from COLORS palette
        country_field: Key in data dicts to use for country summary
        column_widths: Optional dict of column_letter -> width (e.g., {"A": 30})
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    fill_color = COLORS.get(color, color)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # ── Headers ──
    for c, header in enumerate(columns, 1):
        cell = ws.cell(row=1, column=c, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # ── Data rows ──
    col_keys = [col.lower().replace(" ", "_").replace("/", "_") for col in columns]
    for r, row_data in enumerate(data, 2):
        for c, key in enumerate(col_keys, 1):
            val = row_data.get(key, "")
            cell = ws.cell(row=r, column=c, value=clean(val) if isinstance(val, str) else val)
            cell.border = thin_border

    # ── Formatting ──
    ws.freeze_panes = "A2"
    last_col_letter = chr(64 + len(columns)) if len(columns) <= 26 else "Z"
    ws.auto_filter.ref = f"A1:{last_col_letter}{len(data) + 1}"

    if column_widths:
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
    else:
        # Auto-set reasonable widths
        default_widths = [30] + [20] * (len(columns) - 1)
        for i, width in enumerate(default_widths):
            col_letter = chr(65 + i) if i < 26 else chr(65 + i // 26 - 1) + chr(65 + i % 26)
            ws.column_dimensions[col_letter].width = width

    # ── Summary by Country sheet ──
    ws2 = wb.create_sheet("Summary by Country")
    country_counts = Counter(row.get(country_field, "Unknown") for row in data)

    ws2.cell(row=1, column=1, value="Country").font = header_font
    ws2.cell(row=1, column=1).fill = header_fill
    ws2.cell(row=1, column=2, value="Leads").font = header_font
    ws2.cell(row=1, column=2).fill = header_fill

    for r, (country, count) in enumerate(sorted(country_counts.items(), key=lambda x: -x[1]), 2):
        ws2.cell(row=r, column=1, value=clean(country) if isinstance(country, str) else country)
        ws2.cell(row=r, column=2, value=count)

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 12

    wb.save(filename)
    print(f"Saved {filename} — {len(data)} leads from {len(country_counts)} countries")
    return len(data)


# ── Example usage ──
if __name__ == "__main__":
    sample_data = [
        {"name": "Example Business", "country": "Spain", "city": "Barcelona", "website": "https://example.com"},
        {"name": "Test Shop", "country": "Spain", "city": "Madrid", "website": "https://test.com"},
        {"name": "Demo Store", "country": "France", "city": "Paris", "website": "https://demo.fr"},
    ]

    create_leads_xlsx(
        filename="example_leads.xlsx",
        sheet_name="Example Leads",
        columns=["Name", "Country", "City", "Website"],
        data=sample_data,
        color="blue",
        column_widths={"A": 30, "B": 20, "C": 20, "D": 40},
    )
