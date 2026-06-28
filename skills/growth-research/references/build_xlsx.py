"""Build sortable xlsx companions to the research markdown files.

This script is copied into a growth-research workdir as `scripts/build_xlsx.py`.

Reads each `research/<topic>/<topic>.json` (the structured mirror of the
markdown source-of-truth) and writes the matching `research/<topic>/<topic>.xlsx`
with a styled header, frozen pane, autofilter, alternating row shading, and
reasonable column widths.

Run from the repo root:

    python3 scripts/build_xlsx.py

Missing JSON files are skipped with a notice — the script still succeeds.
"""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parents[1]
RESEARCH = REPO / "research"

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(bold=True, color="FFFFFF")
ALT_FILL = PatternFill("solid", fgColor="F3F4F6")

TOPICS = {
    "subreddits": {
        "fields": [
            ("category", "Category"),
            ("subreddit", "Subreddit"),
            ("url", "URL"),
            ("approx_subs", "Approx. subs"),
            ("fit", "Fit"),
            ("geography", "Geography"),
            ("angle", "Angle / use"),
            ("self_promo_flags", "Self-promo / red flags"),
        ],
    },
    "magazines": {
        "fields": [
            ("category", "Category"),
            ("outlet", "Outlet"),
            ("url", "URL"),
            ("country", "Country"),
            ("fit", "Fit"),
            ("audience", "Audience"),
            ("reach", "Reach / size"),
            ("notes", "Notes / contact"),
        ],
    },
    "conferences": {
        "fields": [
            ("category", "Category"),
            ("event", "Event"),
            ("url", "URL"),
            ("dates", "Dates"),
            ("city", "City"),
            ("country", "Country"),
            ("fit", "Fit"),
            ("attendees", "Attendees"),
            ("sponsorship_range", "Sponsorship range"),
            ("why_it_matters", "Why it matters"),
        ],
    },
    "partners": {
        "fields": [
            ("category", "Category"),
            ("company", "Company"),
            ("url", "URL"),
            ("what_they_do", "What they do"),
            ("role", "Role"),
            ("surface", "Surface"),
            ("geography", "Geography"),
            ("pitch", "Pitch angle"),
            ("partner_program_or_contact", "Partner program / contact"),
        ],
    },
    "influencers": {
        "fields": [
            ("category", "Category"),
            ("name", "Name"),
            ("handle", "Handle"),
            ("url", "URL"),
            ("audience_size", "Audience size"),
            ("audience_description", "Audience"),
            ("fit", "Fit"),
            ("geography", "Geography"),
            ("engagement_angle", "Engagement angle"),
            ("contact_path", "Contact path"),
        ],
    },
    "competitor_intel": {
        "fields": [
            ("category", "Category"),
            ("name", "Name"),
            ("url", "URL"),
            ("founded_year", "Founded"),
            ("hq_country", "HQ country"),
            ("employees_approx", "Employees"),
            ("revenue_or_funding", "Revenue / funding"),
            ("pricing_snapshot", "Pricing snapshot"),
            ("key_features", "Key features"),
            ("notable_customers", "Notable customers"),
            ("strengths", "Strengths"),
            ("weaknesses_to_exploit", "Weaknesses to exploit"),
            ("threat_level", "Threat level"),
        ],
    },
    "pricing": {
        "fields": [
            ("category", "Category"),
            ("item", "Item"),
            ("surface", "Surface"),
            ("price_point", "Price point"),
            ("billing_model", "Billing model"),
            ("value_metric", "Value metric"),
            ("target_segment", "Target segment"),
            ("positioning", "Positioning"),
            ("benchmark_anchor", "Benchmark anchor"),
            ("rationale", "Rationale"),
            ("risk_or_note", "Risk / note"),
            ("confidence", "Confidence"),
        ],
    },
    "regulatory": {
        "fields": [
            ("category", "Category"),
            ("jurisdiction", "Jurisdiction"),
            ("regulation_name", "Regulation"),
            ("effective_date", "Effective date"),
            ("status", "Status"),
            ("url", "URL"),
            ("summary", "Summary"),
            ("relevance_to_product", "Relevance"),
            ("compliance_burden_on", "Burden on"),
            ("sales_angle", "Sales angle"),
            ("risk_to_us", "Risk to us"),
            ("notes", "Notes"),
        ],
    },
}


def write_sheet(ws, columns, rows):
    headers = [label for _, label in columns]
    keys = [key for key, _ in columns]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    for i, row in enumerate(rows, start=2):
        ws.append([row.get(k, "") for k in keys])
        if i % 2 == 0:
            for cell in ws[i]:
                cell.fill = ALT_FILL
    ws.freeze_panes = "A2"
    if ws.max_row >= 1:
        ws.auto_filter.ref = ws.dimensions
    # column widths
    widths = [len(h) + 4 for h in headers]
    for row in rows:
        for j, key in enumerate(keys):
            value = row.get(key, "")
            if value:
                widths[j] = max(widths[j], min(60, len(str(value)) + 2))
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build(topic: str, spec: dict) -> int:
    json_path = RESEARCH / topic / f"{topic}.json"
    xlsx_path = RESEARCH / topic / f"{topic}.xlsx"
    if not json_path.exists():
        print(f"{topic}: no JSON found at {json_path.relative_to(REPO)} — skipped")
        return 0
    rows = json.loads(json_path.read_text(encoding="utf-8"))
    if not isinstance(rows, list):
        raise SystemExit(f"{json_path}: expected a JSON array at top level")
    wb = Workbook()
    ws = wb.active
    ws.title = topic
    write_sheet(ws, spec["fields"], rows)
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    print(f"{topic}: {len(rows)} rows → {xlsx_path.relative_to(REPO)}")
    return len(rows)


def main():
    total = 0
    for topic, spec in TOPICS.items():
        total += build(topic, spec)
    print(f"total: {total} rows across {len(TOPICS)} topics")


if __name__ == "__main__":
    main()
